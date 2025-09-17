# mahjong_manager.py
import pandas as pd
import numpy as np
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

class MahjongScoreManager:
    """
    麻雀の成績を管理するクラス。
    チップ有無によるフィルタリング機能と、可変の半荘回数に対応。
    追記時に合計行を動的に移動させ、同日の別メンバー対局にも対応。
    """
    def __init__(self, file_path):
        self.file_path = file_path
        self.raw_data = None
        self.load_data()

    # load_dataは変更ありません
    def load_data(self):
        try:
            xls = pd.ExcelFile(self.file_path)
        except FileNotFoundError:
            self.raw_data = pd.DataFrame()
            return

        all_games_data = []
        sheets_to_process = [s for s in xls.sheet_names if s not in ['result', 'テンプレ']]

        for sheet_name in sheets_to_process:
            try:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                if df.empty or pd.isna(df.iloc[0, 1]):
                    continue

                date_str_clean = re.sub(r'_\d+$', '', sheet_name)
                date_obj = pd.to_datetime(date_str_clean, format='%y%m%d', errors='coerce')
                if pd.isna(date_obj):
                    date_obj = pd.to_datetime(date_str_clean, format='%Y%m%d', errors='coerce')
                if pd.isna(date_obj): continue

                chip_row_index = -1
                if df.shape[1] > 0:
                    chip_rows = df[df.iloc[:, 0].astype(str).str.contains('チップ収支|チップ', na=False, regex=True)].index
                    if not chip_rows.empty:
                        chip_row_index = chip_rows[0]

                game_data_end_row = chip_row_index if chip_row_index != -1 else df.dropna(how='all').index.max() + 1
                chip_data_row = chip_row_index

                players = df.iloc[0, 1:5].tolist()
                game_scores_df = df.iloc[1:game_data_end_row, 1:5]

                chip_series = pd.Series([0,0,0,0])
                if chip_data_row != -1:
                    chip_series = df.iloc[chip_data_row, 1:5].fillna(0)

                has_chips = chip_series.abs().sum() != 0
                chip_flag = 'チップあり' if has_chips else 'チップなし'

                for index, row in game_scores_df.iterrows():
                    if row.isna().all() or pd.isna(row.iloc[0]): continue
                    
                    game_scores = pd.to_numeric(row, errors='coerce').dropna()
                    if len(game_scores) < 4: continue

                    ranks = game_scores.rank(method='min', ascending=False)
                    for i in range(4):
                        all_games_data.append({
                            '日付': date_obj, '名前': players[i],
                            'スコア': game_scores.iloc[i], 'チップ': 0,
                            '順位': ranks.iloc[i], 'チップ有無': chip_flag
                        })
                
                for i in range(4):
                    if chip_series.iloc[i] != 0:
                        all_games_data.append({
                            '日付': date_obj, '名前': players[i],
                            'スコア': 0, 'チップ': chip_series.iloc[i],
                            '順位': np.nan, 'チップ有無': chip_flag
                        })
            except Exception as e:
                print(f"シート '{sheet_name}' の処理中にエラー: {e}")
        
        if all_games_data:
            self.raw_data = pd.DataFrame(all_games_data)
        else:
            self.raw_data = pd.DataFrame()

    def add_record(self, date_str, records):
        """
        マトリックス形式のシートに記録を追記。同日の別メンバーの場合はシートを分けて作成。
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        form_players = [r['名前'] for r in records]
        target_sheet_name = date_str
        
        # --- 既存シートの確認と、追記先シート名の決定ロジック ---
        sheet_to_add_to = None
        
        # まずは日付そのままのシート(例: 250803)があるか確認
        if date_str in wb.sheetnames:
            ws_check = wb[date_str]
            header_players = [ws_check.cell(row=1, column=j).value for j in range(2, 6)]
            if sorted(header_players) == sorted(form_players):
                sheet_to_add_to = date_str

        # 日付+番号のシート(例: 250803_2)を順番に確認
        if not sheet_to_add_to:
            for i in range(2, 20): # _2から_19まで探す
                sheet_name_candidate = f"{date_str}_{i}"
                if sheet_name_candidate in wb.sheetnames:
                    ws_check = wb[sheet_name_candidate]
                    header_players = [ws_check.cell(row=1, column=j).value for j in range(2, 6)]
                    if sorted(header_players) == sorted(form_players):
                        sheet_to_add_to = sheet_name_candidate
                        break
        
        if sheet_to_add_to: # 追記できるシートが見つかった場合
            ws = wb[sheet_to_add_to]
            chip_row_index = -1
            for row in ws.iter_rows(min_row=2, max_col=1):
                cell = row[0]
                if cell.value and ('チップ収支' in str(cell.value) or 'チップ' in str(cell.value)):
                    chip_row_index = cell.row
                    break
            
            if chip_row_index == -1:
                return False, f"エラー: シート '{sheet_to_add_to}' のフォーマットが不正です。「チップ収支」の行が見つかりません。"
            
            new_game_row = chip_row_index
            if (new_game_row - 1) > 20:
                return False, "エラー: このシートは既に対局で満杯です（20半荘まで）。"
            
            ws.insert_rows(new_game_row)

        else: # 追記できるシートがない -> 新規シートを作成
            # 新しいシート名を探す
            if date_str not in wb.sheetnames:
                target_sheet_name = date_str
            else:
                for i in range(2, 20):
                    sheet_name_candidate = f"{date_str}_{i}"
                    if sheet_name_candidate not in wb.sheetnames:
                        target_sheet_name = sheet_name_candidate
                        break
            
            ws = wb.create_sheet(title=target_sheet_name)
            for i, name in enumerate(form_players):
                ws.cell(row=1, column=i+2, value=name)
            new_game_row = 2

        # --- 書き込み共通処理 ---
        scores = [r['SCORE'] for r in records]
        chips = [r['チップ'] for r in records]
        ranks = pd.Series(scores).rank(method='min', ascending=False).astype(int).tolist()

        ws.cell(row=new_game_row, column=1).value = new_game_row - 1
        for i in range(4):
            ws.cell(row=new_game_row, column=i+2, value=scores[i])
            ws.cell(row=new_game_row, column=i+9, value=ranks[i])

        chip_balance_row = new_game_row + 1
        total_score_row = new_game_row + 2
        
        ws.cell(row=chip_balance_row, column=1).value = 'チップ収支'
        ws.cell(row=total_score_row, column=1).value = '合計'

        for i in range(4):
            prev_chip_val = ws.cell(row=chip_balance_row, column=i+2).value or 0
            prev_total_val = ws.cell(row=total_score_row, column=i+2).value or 0
            ws.cell(row=chip_balance_row, column=i+2).value = prev_chip_val + chips[i]
            ws.cell(row=total_score_row, column=i+2).value = prev_total_val + scores[i]
            
        wb.save(self.file_path)
        self.load_data()
        return True, f"シート'{target_sheet_name}'の{new_game_row - 1}半荘目として記録を追記しました。"

    def get_player_list(self):
        if self.raw_data is None or self.raw_data.empty: return []
        return sorted(self.raw_data['名前'].unique())

    def calculate_stats(self, start_date=None, end_date=None, chip_filter='全て'):
        if self.raw_data is None or self.raw_data.empty: return pd.DataFrame()
        
        df_filtered = self.raw_data.copy()

        if chip_filter == 'チップありのみ':
            df_filtered = df_filtered[df_filtered['チップ有無'] == 'チップあり']
        elif chip_filter == 'チップなしのみ':
            df_filtered = df_filtered[df_filtered['チップ有無'] == 'チップなし']
        
        if start_date:
            df_filtered = df_filtered[df_filtered['日付'].dt.date >= start_date]
        if end_date:
            df_filtered = df_filtered[df_filtered['日付'].dt.date <= end_date]
        
        if df_filtered.empty: return pd.DataFrame()

        game_counts = df_filtered[df_filtered['スコア'] != 0].groupby('名前')['名前'].count()
        
        df_filtered['トータルスコア'] = df_filtered['スコア'] + df_filtered['チップ']
        stats = df_filtered.groupby('名前').agg(
            トータルスコア=('トータルスコア', 'sum'),
            チップスコア=('チップ', 'sum'),
            平均着順=('順位', 'mean')
        ).reset_index()

        stats = pd.merge(stats, game_counts.rename('対戦回数'), on='名前', how='left').fillna(0)
        stats['平均着順'] = stats['平均着順'].round(2)
        
        return stats